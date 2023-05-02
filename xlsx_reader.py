from openpyxl import load_workbook
from constants import CLASSES


def get_class_schedule(number, letter):
    """return schedule by class from school_schedule.xlsx document"""
    wb = load_workbook("static/school_schedule.xlsx")
    sheet = wb.get_sheet_by_name(f"{number} класс")
    index = CLASSES.index(letter)
    table = []
    for i in range(7 * index + 1, 7 * (index + 1)):
        row = []
        for j in range(2, 8):
            value = sheet.cell(row=i, column=j).value
            if not value:
                break
            row.append(value)
        if row:
            table.append(row)
    return table
